import os, time, json, random, io
from datetime import datetime, timedelta
from flask import Blueprint, request, jsonify, g, send_file
import pandas as pd
from werkzeug.security import generate_password_hash, check_password_hash
from database.connection import connecter_base_de_donnees, rows_to_dicts
from utils.globals import JWT_SECRET, JWT_EXPIRATION_HOURS, ADMIN_LOGS, SYSTEM_SETTINGS
from auth.decorators import token_required, admin_required, role_required
from ia.prediction import run_ai_prediction, get_ia_report_anomalies
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm


reports_bp = Blueprint('reports_bp', __name__)

@reports_bp.route("/reports/history", methods=["GET"])
@token_required
def get_reports_history():
    """
    [PFE] Génère dynamiquement la liste des rapports disponibles
    basée sur la date d'aujourd'hui.
    """
    import datetime
    today_str = datetime.datetime.now().strftime("%d/%m/%Y")
    return jsonify([
        {"id": "REP-DAILY-01", "nom": "Rapport Quotidien NOC", "type": "daily", "date": today_str + " 08:00", "auteur": "Système", "taille": "2.4 MB", "statut": "Prêt"},
        {"id": "REP-IA-02", "nom": "Analyse Anomalies IA", "type": "ia", "date": today_str + " 10:30", "auteur": "IA Auto", "taille": "1.1 MB", "statut": "Prêt"}
    ])

def create_pdf_canvas(title):
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    c.setFont("Helvetica-Bold", 16)
    c.drawString(2*cm, height - 2*cm, title)
    c.setFont("Helvetica", 10)
    c.drawString(2*cm, height - 2.8*cm, f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    c.drawString(2*cm, height - 3.2*cm, "Tunisie Telecom - Direction Régionale Mahdia")
    c.line(2*cm, height - 3.5*cm, width - 2*cm, height - 3.5*cm)
    return buffer, c, width, height

@reports_bp.route("/reports/daily-pdf", methods=["GET"])
@token_required
def generate_daily_pdf():
    try:
        conn = connecter_base_de_donnees()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM antennes")
        total = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM antennes_statut WHERE statut != 'normal'")
        alertes = cur.fetchone()[0]
        cur.execute("SELECT AVG(disponibilite) FROM antennes_statut")
        dispo = cur.fetchone()[0]
        cur.execute("SELECT * FROM incidents WHERE statut != 'resolu' ORDER BY date_creation DESC LIMIT 10")
        incidents = rows_to_dicts(cur)
        cur.close()
        conn.close()

        buffer, c, w, h = create_pdf_canvas("RAPPORT QUOTIDIEN NOC")
        
        c.setFont("Helvetica-Bold", 12)
        c.drawString(2*cm, h - 4.5*cm, "Résumé de l'état du réseau:")
        c.setFont("Helvetica", 11)
        c.drawString(2.5*cm, h - 5.2*cm, f"Total antennes supervisées : {total}")
        c.drawString(2.5*cm, h - 5.8*cm, f"Antennes en alerte/critique : {alertes}")
        c.drawString(2.5*cm, h - 6.4*cm, f"Disponibilité globale : {dispo:.2f}%")
        
        c.setFont("Helvetica-Bold", 12)
        c.drawString(2*cm, h - 8*cm, "Incidents actifs:")
        
        y = h - 8.8*cm
        c.setFont("Helvetica", 10)
        for inc in incidents:
            c.drawString(2.5*cm, y, f"- {inc['titre']} (Antenne {inc['antenne_id']}) - Criticité: {inc['criticite']}")
            y -= 0.6*cm
            if y < 2*cm:
                c.showPage()
                y = h - 2*cm

        c.save()
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name="rapport_quotidien.pdf", mimetype="application/pdf")
    except Exception as e:
        print("Erreur PDF:", e)
        return jsonify({"error": str(e)}), 500

@reports_bp.route("/rapport/quotidien", methods=["GET"])
@token_required
def rapport_quotidien():
    """Rapport PDF quotidien — état global du réseau."""
    try:
        conn = connecter_base_de_donnees()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM antennes")
        total = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM antennes_statut WHERE statut IN ('alerte', 'critique')")
        alertes = cur.fetchone()[0]
        cur.execute("SELECT AVG(disponibilite), AVG(cpu), AVG(temperature) FROM antennes_statut")
        row = cur.fetchone()
        dispo, avg_cpu, avg_temp = (row[0] or 100), (row[1] or 0), (row[2] or 0)
        cur.execute("SELECT titre, criticite, statut, date_creation FROM incidents WHERE statut != 'resolu' ORDER BY date_creation DESC LIMIT 10")
        incidents = cur.fetchall()
        cur.close(); conn.close()

        buffer, c, w, h = create_pdf_canvas("RAPPORT QUOTIDIEN NOC — Tunisie Télécom Mahdia")
        y = h - 5.0*cm
        c.setFont("Helvetica-Bold", 12)
        c.drawString(2*cm, y, "Résumé de l'état du réseau")
        y -= 0.8*cm
        c.setFont("Helvetica", 11)
        for line in [
            f"Total antennes supervisées : {total}",
            f"Antennes en alerte ou critique : {alertes}",
            f"Disponibilité globale : {dispo:.2f}%",
            f"Charge CPU moyenne : {avg_cpu:.1f}%",
            f"Température moyenne : {avg_temp:.1f}°C",
        ]:
            c.drawString(2.5*cm, y, line)
            y -= 0.6*cm

        y -= 0.4*cm
        c.setFont("Helvetica-Bold", 12)
        c.drawString(2*cm, y, "Incidents actifs")
        y -= 0.8*cm
        c.setFont("Helvetica", 10)
        if not incidents:
            c.drawString(2.5*cm, y, "Aucun incident actif au moment de la génération du rapport.")
        for inc in incidents:
            c.drawString(2.5*cm, y, f"• {inc[0]}  [{inc[1]}]  — Détecté le {str(inc[3])[:16]}")
            y -= 0.6*cm
            if y < 3*cm: c.showPage(); y = h - 2*cm
        c.save()
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name="rapport_quotidien.pdf", mimetype="application/pdf")
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@reports_bp.route("/rapport/incidents", methods=["GET"])
@token_required
def rapport_incidents():
    """Rapport PDF — liste complète des incidents."""
    try:
        conn = connecter_base_de_donnees()
        cur = conn.cursor()
        cur.execute("""
            SELECT i.titre, i.criticite, i.statut, i.description, i.date_creation, a.nom, a.zone
            FROM incidents i JOIN antennes a ON i.antenne_id = a.id
            ORDER BY i.date_creation DESC LIMIT 50
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()

        buffer, c, w, h = create_pdf_canvas("RAPPORT DES INCIDENTS — Isolation Forest")
        y = h - 5.0*cm
        c.setFont("Helvetica", 10)
        for row in rows:
            titre, crit, statut, desc, date, nom, zone = row
            c.setFont("Helvetica-Bold", 10)
            c.drawString(2*cm, y, f"• {titre}  [{crit.upper()}]  — {nom} ({zone})")
            y -= 0.5*cm
            c.setFont("Helvetica", 9)
            c.drawString(2.5*cm, y, f"{str(date)[:16]}  |  Statut : {statut}")
            y -= 0.5*cm
            if desc:
                c.drawString(2.5*cm, y, (desc[:90] + "…") if len(desc) > 90 else desc)
                y -= 0.5*cm
            y -= 0.3*cm
            if y < 3*cm: c.showPage(); y = h - 2*cm
        c.save()
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name="rapport_incidents.pdf", mimetype="application/pdf")
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@reports_bp.route("/rapport/ia", methods=["GET"])
@token_required
def rapport_ia():
    """Rapport PDF — analyse Isolation Forest."""
    try:
        anomalies, total_analysed = get_ia_report_anomalies()

        buffer, c, w, h = create_pdf_canvas("RAPPORT ANALYSE IA — Isolation Forest")
        y = h - 5.0*cm
        c.setFont("Helvetica-Bold", 12)
        c.drawString(2*cm, y, f"Anomalies détectées : {len(anomalies)} / {total_analysed} sites analysés")
        y -= 1.0*cm
        c.setFont("Helvetica", 10)
        for _, row in anomalies.head(20).iterrows():
            c.setFont("Helvetica-Bold", 10)
            c.drawString(2*cm, y, f"• {row['nom']}  ({row['zone']})  —  Score IA : {row['score']:.1f}%")
            y -= 0.5*cm
            c.setFont("Helvetica", 9)
            c.drawString(2.5*cm, y, f"CPU={row['cpu']:.1f}%  Temp={row['temperature']:.1f}°C  Latence={row['latence']:.0f}ms  Dispo={row['disponibilite']:.1f}%")
            y -= 0.7*cm
            if y < 3*cm: c.showPage(); y = h - 2*cm
        c.save()
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name="rapport_ia.pdf", mimetype="application/pdf")
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@reports_bp.route("/rapport/antennes/excel", methods=["GET"])
@token_required
def rapport_antennes_excel():
    """Export Excel — données techniques des 127 antennes."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        conn = connecter_base_de_donnees()
        df = pd.read_sql("""
            SELECT id, nom, zone, type, temperature, cpu, ram, signal,
                   latence, packet_loss, disponibilite, debit, statut, date_mesure::text
            FROM antennes_statut ORDER BY id
        """, conn)
        conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Antennes Mahdia"
        headers = ["ID", "Nom", "Zone", "Type", "Temp (°C)", "CPU (%)", "RAM (%)",
                   "Signal (dBm)", "Latence (ms)", "Perte Paquets (%)", "Dispo (%)", "Débit (Mbps)", "Statut", "Date Mesure"]
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = max(12, len(h) + 2)
        for r_idx, row in df.iterrows():
            ws.append(list(row))

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name="antennes_export.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@reports_bp.route("/reports/csv-data", methods=["GET"])
@token_required
def export_csv_data():
    export_type = request.args.get("type", "antennes")
    try:
        conn = connecter_base_de_donnees()
        if export_type == "antennes":
            df = pd.read_sql("SELECT id, nom, zone, type, temperature, cpu, signal, packet_loss, disponibilite, statut FROM antennes_statut", conn)
        elif export_type == "incidents":
            df = pd.read_sql("SELECT * FROM incidents ORDER BY date_creation DESC", conn)
        else:
            df = pd.read_sql("SELECT * FROM mesures ORDER BY date_mesure DESC LIMIT 500", conn)
        conn.close()
        
        output = io.StringIO()
        df.to_csv(output, index=False)
        output.seek(0)
        
        mem = io.BytesIO()
        mem.write(output.getvalue().encode('utf-8'))
        mem.seek(0)
        
        return send_file(mem, mimetype="text/csv", as_attachment=True, download_name=f"{export_type}.csv")
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@reports_bp.route("/api/mesures/live", methods=["GET"])
@token_required
def get_mesures_live():
    # Return latest measures for live charts
    conn = connecter_base_de_donnees()
    cur = conn.cursor()
    cur.execute("""
        SELECT m.*, a.nom as antenne_nom 
        FROM mesures m 
        JOIN antennes a ON m.antenne_id = a.id 
        ORDER BY m.date_mesure DESC LIMIT 50
    """)
    data = rows_to_dicts(cur)
    cur.close()
    conn.close()
    return jsonify(data)

@reports_bp.route("/api/test-ia", methods=["POST"])
@token_required
def force_ia_incident():
    """
    Injecte des métriques réseau extrêmes pour un scénario de test.
    Le statut est ensuite déterminé EXCLUSIVEMENT par l'Isolation Forest.
    Ne code aucun statut en dur.
    """
    data          = request.json or {}
    type_incident = data.get("type", "surchauffe")

    # Métriques de base (réseau sain)
    metrics = {
        "temp": 38.0, "cpu": 35.0, "signal": -62.0,
        "traffic": 120.0, "ram": 45.0, "latence": 18.0,
        "dispo": 99.5, "packet_loss": 0.2, "jitter": 4.0
    }

    # Injection de métriques extrêmes selon le scénario
    if type_incident == "surchauffe":
        metrics["temp"] = 91.0
        metrics["cpu"]  = 96.0
    elif type_incident == "surcharge":
        metrics["cpu"]     = 99.0
        metrics["traffic"] = 950.0
        metrics["ram"]     = 97.0
        metrics["latence"] = 280.0
    elif type_incident == "panne":
        metrics["dispo"]      = 55.0
        metrics["signal"]     = -118.0
        metrics["traffic"]    = 0.0
        metrics["packet_loss"]= 18.0

    try:
        conn = connecter_base_de_donnees()
        cur  = conn.cursor()
        cur.execute("SELECT id FROM antennes ORDER BY RANDOM() LIMIT 1")
        ant_id = cur.fetchone()[0]

        # Statut = 'analyse_ia_en_cours' — l'IA décidera après
        cur.execute("""
            INSERT INTO mesures
                (antenne_id, temperature, cpu, signal, traffic, ram,
                 latence, disponibilite, packet_loss, jitter, statut, date_mesure)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'analyse_ia_en_cours', NOW())
        """, (
            ant_id,
            metrics["temp"], metrics["cpu"], metrics["signal"], metrics["traffic"],
            metrics["ram"],  metrics["latence"], metrics["dispo"],
            metrics["packet_loss"], metrics["jitter"]
        ))
        conn.commit()
        cur.close()
        conn.close()

        # Déclenche l'analyse IA immédiatement pour que le statut soit mis à jour
        import requests as _req
        _req.get("http://localhost:5000/internal/predict", timeout=15)

        return jsonify({"success": True,
                        "message": f"Scénario '{type_incident}' injecté sur l'antenne {ant_id}. L'IA analyse en cours."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

